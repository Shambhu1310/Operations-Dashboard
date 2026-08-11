import openpyxl, os, math, json
from collections import defaultdict

gl_dir = r'c:\Users\shash\OneDrive\Desktop\Term 2\OpsLift\data\great_lakes'

files_meta = [
    {'batch': 'PGPM', 'total_students': 130, 'file': 'PGPM-27_Term-2_Time_Table.xlsx'},
    {'batch': 'PGDM 2', 'total_students': 300, 'file': 'PGDM_2_Term_4_Timetable.xlsx'},
    {'batch': 'PGDM 1', 'total_students': 360, 'file': 'PGDM_1_Term_1_Timetable.xlsx'}
]

all_classes = []
batch_sections = {}
batch_strengths = {}

for item in files_meta:
    batch = item['batch']
    p = os.path.join(gl_dir, item['file'])
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    
    sec_col = headers.index('Section')
    sections = sorted(list(set(r[sec_col] for r in rows if r[sec_col])))
    batch_sections[batch] = sections
    sec_count = len(sections)
    students_per_sec = item['total_students'] / sec_count
    batch_strengths[batch] = students_per_sec
    
    day_col = headers.index('Day')
    date_col = headers.index('Date')
    start_col = headers.index('Start_Time')
    end_col = headers.index('End_Time')
    course_code_col = headers.index('Course_Code')
    course_name_col = headers.index('Course_Name')
    faculty_col = headers.index('Faculty')
    room_col = headers.index('Room_Code')
    
    for r in rows:
        room = str(r[room_col]).strip() if r[room_col] else ''
        all_classes.append({
            'batch': batch,
            'day': r[day_col],
            'date': str(r[date_col]),
            'start_time': str(r[start_col]),
            'end_time': str(r[end_col]),
            'section': r[sec_col],
            'course_code': r[course_code_col],
            'course_name': r[course_name_col],
            'faculty': r[faculty_col],
            'room': room,
            'total_batch_students': item['total_students'],
            'students_per_sec': students_per_sec
        })

print('=== 1. BATCH & SECTION STATS ===')
for b in batch_sections:
    print(f'{b}: {len(batch_sections[b])} sections {batch_sections[b]} -> Avg {batch_strengths[b]:.1f} students/section')

# Filter for Academic Block A
a_block_classes = []
non_a_classes = []
a_rooms = set()
floors = set()

for c in all_classes:
    room = c['room']
    if room.startswith('A') and len(room) >= 2 and room[1].isdigit():
        floor = int(room[1])
        c['is_a_block'] = True
        c['floor'] = floor
        a_rooms.add(room)
        floors.add(floor)
        a_block_classes.append(c)
    else:
        c['is_a_block'] = False
        non_a_classes.append(c)

print(f'\nTotal classes across all timetables: {len(all_classes)}')
print(f'Academic Block A classes: {len(a_block_classes)}')
print(f'Non-A Block classes: {len(non_a_classes)}')
print(f'A-Block Rooms: {sorted(list(a_rooms))}')
print(f'Floors identified: {sorted(list(floors))}')

# Build Events: Class Start (Up), Class End (Down)
events = []
for c in a_block_classes:
    students = c['students_per_sec']
    u15 = students * 0.15
    u25 = students * 0.25
    u35 = students * 0.35
    events.append({
        'batch': c['batch'],
        'day': c['day'],
        'date': c['date'],
        'time': c['start_time'],
        'direction': 'Up',
        'event_type': 'Class Start - Up',
        'section': c['section'],
        'course_name': c['course_name'],
        'room': c['room'],
        'floor': c['floor'],
        'students': students,
        'users_15': u15,
        'users_25': u25,
        'users_35': u35,
        'trips_15_cap7': math.ceil(u15 / 7),
        'trips_25_cap7': math.ceil(u25 / 7),
        'trips_35_cap7': math.ceil(u35 / 7),
    })
    events.append({
        'batch': c['batch'],
        'day': c['day'],
        'date': c['date'],
        'time': c['end_time'],
        'direction': 'Down',
        'event_type': 'Class End - Down',
        'section': c['section'],
        'course_name': c['course_name'],
        'room': c['room'],
        'floor': c['floor'],
        'students': students,
        'users_15': u15,
        'users_25': u25,
        'users_35': u35,
        'trips_15_cap7': math.ceil(u15 / 7),
        'trips_25_cap7': math.ceil(u25 / 7),
        'trips_35_cap7': math.ceil(u35 / 7),
    })

print(f'\nTotal Timetable Events (Up + Down): {len(events)}')

# Aggregation by (Day, Date, Time)
day_time_agg = defaultdict(lambda: {
    'up_users_15': 0.0, 'up_users_25': 0.0, 'up_users_35': 0.0,
    'down_users_15': 0.0, 'down_users_25': 0.0, 'down_users_35': 0.0,
    'total_users_15': 0.0, 'total_users_25': 0.0, 'total_users_35': 0.0,
    'sections_up': [], 'sections_down': [],
    'floors': set(), 'batches': set()
})

for e in events:
    key = (e['day'], e['date'], e['time'])
    agg = day_time_agg[key]
    agg['batches'].add(e['batch'])
    agg['floors'].add(e['floor'])
    
    if e['direction'] == 'Up':
        agg['up_users_15'] += e['users_15']
        agg['up_users_25'] += e['users_25']
        agg['up_users_35'] += e['users_35']
        agg['sections_up'].append(f"{e['batch']} {e['section']} ({e['room']})")
    else:
        agg['down_users_15'] += e['users_15']
        agg['down_users_25'] += e['users_25']
        agg['down_users_35'] += e['users_35']
        agg['sections_down'].append(f"{e['batch']} {e['section']} ({e['room']})")
    
    agg['total_users_15'] += e['users_15']
    agg['total_users_25'] += e['users_25']
    agg['total_users_35'] += e['users_35']

# Peak analysis for Monday
monday_aggs = [(k, v) for k, v in day_time_agg.items() if k[0] == 'Monday']
monday_aggs_sorted = sorted(monday_aggs, key=lambda x: x[1]['total_users_25'], reverse=True)

print('\n=== PEAK EVENTS ON MONDAY (25% Base Scenario) ===')
for (day, date, time), v in monday_aggs_sorted[:8]:
    trips_cap7 = math.ceil(v['total_users_25'] / 7)
    trips_cap6 = math.ceil(v['total_users_25'] / 6)
    trips_cap8 = math.ceil(v['total_users_25'] / 8)
    conc_count = len(v['sections_up']) + len(v['sections_down'])
    print(f"{day} {time} | Up: {v['up_users_25']:.1f} pax, Down: {v['down_users_25']:.1f} pax | Total: {v['total_users_25']:.1f} (~{round(v['total_users_25'])} pax) | Trips (Cap 7): {trips_cap7} (Cap 6: {trips_cap6}, Cap 8: {trips_cap8}) | Concurrent: {conc_count} sections")

# Floor Demand
floor_demand_25 = defaultdict(float)
for e in events:
    floor_demand_25[e['floor']] += e['users_25']
print('\n=== TOTAL DEMAND BY FLOOR (25% Base Scenario) ===')
for fl, d in sorted(floor_demand_25.items(), key=lambda x: x[1], reverse=True):
    print(f"Floor {fl}: {d:.1f} pax ({round(d)} pax, {d/sum(floor_demand_25.values())*100:.1f}%)")

# Batch Demand
batch_demand_25 = defaultdict(float)
for e in events:
    batch_demand_25[e['batch']] += e['users_25']
print('\n=== TOTAL DEMAND BY PROGRAMME (25% Base Scenario) ===')
for b, d in sorted(batch_demand_25.items(), key=lambda x: x[1], reverse=True):
    print(f"{b}: {d:.1f} pax ({round(d)} pax, {d/sum(batch_demand_25.values())*100:.1f}%)")

# Export to clean JSON for web application integration
export_data = {
    'metadata': {
        'building': 'Great Lakes Academic Block A',
        'elevators': 1,
        'floors': [1, 2, 3],
        'observed_capacity_range': [6, 7, 8],
        'base_capacity': 7,
        'scenarios': [0.15, 0.25, 0.35],
        'batches': [
            {'batch': 'PGPM', 'students': 130, 'sections': batch_sections['PGPM'], 'students_per_sec': batch_strengths['PGPM']},
            {'batch': 'PGDM 2', 'students': 300, 'sections': batch_sections['PGDM 2'], 'students_per_sec': batch_strengths['PGDM 2']},
            {'batch': 'PGDM 1', 'students': 360, 'sections': batch_sections['PGDM 1'], 'students_per_sec': batch_strengths['PGDM 1']}
        ],
        'a_rooms': sorted(list(a_rooms)),
        'floors_list': sorted(list(floors))
    },
    'classes': a_block_classes,
    'events': events
}

out_json = r'c:\Users\shash\OneDrive\Desktop\Term 2\OpsLift\js\great_lakes_data.js'
with open(out_json, 'w', encoding='utf-8') as f:
    f.write('const GREAT_LAKES_DATA = ' + json.dumps(export_data, indent=2) + ';\n')
print(f'\nWrote Great Lakes data to {out_json}')
